#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
QA用表统计生成脚本
通过RAGFlow智能体工作流调用LLM提取查询条件和查询内容
生成17列格式的用表统计Excel文件

使用方法：
1. 在RAGFlow中创建工作流，配置Agent节点（参见workflow_config.md）
2. 将Agent ID填入下方CONFIG配置
3. 运行脚本: python generate_table_statistics.py
"""

import pandas as pd
import requests
import json
import re
import time
import logging
from pathlib import Path
from collections import defaultdict
from itertools import groupby
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from typing import List, Dict, Tuple, Optional

# ============================================================================
# 配置区 - 请根据实际情况修改
# ============================================================================

CONFIG = {
    # RAGFlow API配置
    "base_url": "http://172.17.2.65:80",  # RAGFlow服务地址
    "api_key": "ragflow-SSwrhX-_RYND1jbNEihxw5fGevFQHRQwdTXXiYund38",  # API密钥
    
    # 智能体ID - 从RAGFlow工作流中复制（创建工作流后填入）
    "agent_id": "6d6e5350f67911f0ab7e0242ad420006",  # <<<< 请替换为实际的Agent ID
    
    # 请求配置
    "timeout": 60,           # 请求超时时间（秒）
    "retry_count": 3,        # 重试次数
    "retry_delay": 2,        # 重试间隔（秒）
    "request_delay": 0.5,    # 请求间隔（秒），避免请求过快
}

# 系统名称映射
SYSTEM_MAPPING = {
    '机电': '高精度地图可视化_机电一张图',
    '路运': '路段路运一体化',
    '原始': '数据资源管理平台',
    '报表': '报表台账',
}

# 数据库名映射配置（表名前缀 -> 数据库名）
DATABASE_MAPPING = {
    'darm_': 'darm',           # 报表台账表
    'gjdt_': 'datarms_ods',    # 机电数据表
    'ldly_': 'datarms_ods',    # 路运数据表
    'jtsjzy_': 'datarms_ods',  # 原始表数据
    'dim_': 'datarms_ods',     # 维度表
}
DEFAULT_DATABASE = 'datarms_ods'

# 输入文件配置
INPUT_FILES = [
    ('../新博数据管家_问数_Q_SQL_v3.0.1/q_sql/新博报表q_sql_v3.0.3.xlsx', '报表'),
    ('../新博数据管家_问数_Q_SQL_v3.0.1/q_sql/新博机电数据q_sql_v3.0.4.xlsx', '机电'),
    ('../新博数据管家_问数_Q_SQL_v3.0.1/q_sql/新博路运q_sql_v2.0.3.xlsx', '路运'),
    ('../新博数据管家_问数_Q_SQL_v3.0.1/q_sql/新博原始表q_sql_v3.0.3.xlsx', '原始'),
]

# 输出配置
OUTPUT_DIR = Path('sortqa')
OUTPUT_FILE = OUTPUT_DIR / '用表统计_生成.xlsx'

# ============================================================================
# 日志配置
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# 核心类
# ============================================================================

class QAClassificationExtractor:
    """通过RAGFlow工作流提取查询条件和查询内容"""
    
    def __init__(self, base_url: str, api_key: str, agent_id: str):
        self.base_url = base_url.rstrip('/')
        self.api_key = api_key
        self.agent_id = agent_id
        self.api_url = f"{self.base_url}/api/v1/agents/{agent_id}/completions"
        
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        })
        
        # 缓存已提取的结果
        self.cache = {}
    
    def extract(self, question: str) -> Tuple[str, str]:
        """
        调用工作流提取查询条件和查询内容
        
        Args:
            question: 用户问题
            
        Returns:
            (查询条件, 查询内容)
        """
        if not question or pd.isna(question):
            return "", ""
        
        question = str(question).strip()
        
        # 检查缓存
        if question in self.cache:
            return self.cache[question]
        
        payload = {
            "question": question,
            "stream": False
        }
        
        for attempt in range(CONFIG["retry_count"]):
            try:
                response = self.session.post(
                    self.api_url,
                    json=payload,
                    timeout=CONFIG["timeout"]
                )
                
                if response.status_code != 200:
                    logger.warning(f"API请求失败，状态码: {response.status_code}")
                    continue
                
                data = response.json()
                result = self._parse_response(data)
                
                if result:
                    self.cache[question] = result
                    return result
                    
            except requests.exceptions.Timeout:
                logger.warning(f"请求超时 - 尝试 {attempt + 1}/{CONFIG['retry_count']}")
            except Exception as e:
                logger.error(f"请求异常: {e}")
            
            if attempt < CONFIG["retry_count"] - 1:
                time.sleep(CONFIG["retry_delay"])
        
        # 失败时使用正则兜底
        result = self._regex_fallback(question)
        self.cache[question] = result
        return result
    
    def _parse_response(self, data: dict) -> Optional[Tuple[str, str]]:
        """解析API响应，提取查询条件和查询内容"""
        try:
            if data.get("code") != 0:
                return None
            
            # 尝试多种路径获取content
            content = None
            
            if "data" in data:
                resp_data = data["data"]
                if isinstance(resp_data, dict):
                    if "data" in resp_data and isinstance(resp_data["data"], dict):
                        content = resp_data["data"].get("content", "")
                    elif "content" in resp_data:
                        content = resp_data["content"]
            
            if not content:
                return None
            
            # 解析JSON输出
            # 尝试提取JSON
            json_match = re.search(r'\{[^{}]*"查询条件"[^{}]*\}', content, re.DOTALL)
            if json_match:
                try:
                    result = json.loads(json_match.group())
                    condition = result.get("查询条件", "")
                    query_content = result.get("查询内容", "")
                    return (condition, query_content)
                except json.JSONDecodeError:
                    pass
            
            # 尝试从文本中提取
            condition_match = re.search(r'查询条件[：:]\s*(.+?)(?:\n|查询内容|$)', content)
            content_match = re.search(r'查询内容[：:]\s*(.+?)(?:\n|$)', content)
            
            if condition_match or content_match:
                condition = condition_match.group(1).strip() if condition_match else ""
                query_content = content_match.group(1).strip() if content_match else ""
                return (condition, query_content)
            
            return None
            
        except Exception as e:
            logger.error(f"解析响应失败: {e}")
            return None
    
    def _regex_fallback(self, question: str) -> Tuple[str, str]:
        """正则兜底提取"""
        condition = ""
        content = ""
        
        # 模式1: "查询/统计 所有xxx 的 yyy"
        m = re.match(r'^(?:查询|统计|获取)(?:所有|全部)(.+?)的(.+?)$', question)
        if m:
            return (f"所有{m.group(1).strip()}", m.group(2).strip())
        
        # 模式2: "查询/统计 xxx为yyy 的 zzz"
        m = re.match(r'^(?:查询|统计)(.+?(?:为|是|等于).+?)的(.+?)$', question)
        if m:
            return (m.group(1).strip(), m.group(2).strip())
        
        # 模式3: "查询所有xxx信息"
        m = re.match(r'^(?:查询|统计)(?:所有|全部)(.+?)(?:信息|数据)?$', question)
        if m:
            content = m.group(1).strip()
            if not content.endswith('信息'):
                content += "信息"
            return ("所有", content)
        
        # 模式4: "按xxx统计yyy"
        m = re.match(r'^按(.+?)(?:统计|分组)(.+?)$', question)
        if m:
            return (f"按{m.group(1).strip()}", m.group(2).strip())
        
        # 模式5: "统计xxx"
        m = re.match(r'^统计(.+?)$', question)
        if m:
            return ("", m.group(1).strip())
        
        # 模式6: "查询xxx"
        m = re.match(r'^(?:查询|获取)(.+?)$', question)
        if m:
            raw = m.group(1).strip()
            # 检查是否有"的"分割
            if '的' in raw:
                parts = raw.rsplit('的', 1)
                if len(parts) == 2:
                    return (parts[0].strip(), parts[1].strip())
            return ("", raw)
        
        return ("", question)


class TableStatisticsGenerator:
    """用表统计生成器"""
    
    def __init__(self, extractor: QAClassificationExtractor):
        self.extractor = extractor
    
    def extract_tables(self, sql: str) -> List[str]:
        """从SQL中提取表名"""
        if pd.isna(sql) or not sql:
            return []
        
        # 支持反引号包裹的表名
        from_pattern = r'FROM\s+`?([a-zA-Z_][a-zA-Z0-9_]*)`?(?:\s+(?:AS\s+)?[a-zA-Z_]|\s*$|\s*,|\s+WHERE|\s+LEFT|\s+RIGHT|\s+INNER|\s+JOIN|\s+ORDER|\s+GROUP|\s+LIMIT|\s*\))'
        join_pattern = r'(?:LEFT|RIGHT|INNER|OUTER|CROSS)?\s*JOIN\s+`?([a-zA-Z_][a-zA-Z0-9_]*)`?(?:\s+(?:AS\s+)?[a-zA-Z_]|\s*$|\s+ON)'
        
        tables = re.findall(from_pattern, sql, re.IGNORECASE)
        join_tables = re.findall(join_pattern, sql, re.IGNORECASE)
        tables.extend(join_tables)
        
        if not tables:
            simple_from = r'FROM\s+`?([a-zA-Z_][a-zA-Z0-9_]*)`?'
            tables = re.findall(simple_from, sql, re.IGNORECASE)
        
        seen = set()
        result = []
        for t in tables:
            t_lower = t.lower()
            if t_lower not in seen:
                seen.add(t_lower)
                result.append(t_lower)
        
        return result
    
    def get_database_name(self, table_name: str) -> str:
        """根据表名获取数据库名"""
        table_lower = table_name.lower()
        for prefix, db_name in DATABASE_MAPPING.items():
            if table_lower.startswith(prefix):
                return db_name
        return DEFAULT_DATABASE
    
    def process_qa_file(self, file_path: str, system_key: str) -> List[Dict]:
        """处理单个QA文件"""
        system_name = SYSTEM_MAPPING.get(system_key, '未知系统')
        logger.info(f"\n处理文件: {file_path}")
        logger.info(f"所属系统: {system_name}")
        
        try:
            df = pd.read_excel(file_path, header=None)
            df.columns = ['问题', 'SQL']
            
            data = []
            total = len(df)
            
            for idx, row in df.iterrows():
                question = row['问题']
                sql = row['SQL']
                tables = self.extract_tables(sql)
                
                if not tables:
                    continue
                
                main_table = tables[0]
                main_db = self.get_database_name(main_table)
                
                # 判断查询类型
                is_multi_table = len(tables) > 1
                related_tables = tables[1:] if is_multi_table else []
                related_dbs = list(set([self.get_database_name(t) for t in related_tables]))
                is_cross_db = len(set([main_db] + related_dbs)) > 1
                
                # 调用LLM提取查询条件和内容
                condition, content = self.extractor.extract(question)
                
                record = {
                    '所属系统': system_name,
                    '所属数据库': main_db,
                    '表名': main_table,
                    '判断条件': condition,
                    '查询内容': content,
                    '单表查询入库问题': '' if (is_multi_table or is_cross_db) else question,
                    '单表查询入库sql': '' if (is_multi_table or is_cross_db) else sql,
                    '跨表查询关联表格': ', '.join(related_tables) if (is_multi_table and not is_cross_db) else '',
                    '跨表查询入库问题': question if (is_multi_table and not is_cross_db) else '',
                    '跨表查询入库sql': sql if (is_multi_table and not is_cross_db) else '',
                    '跨库查询关联数据库': ', '.join(related_dbs) if is_cross_db else '',
                    '跨库查询关联表格': ', '.join(related_tables) if is_cross_db else '',
                    '跨库查询入库问题': question if is_cross_db else '',
                    '跨库查询入库sql': sql if is_cross_db else '',
                    '_is_multi_table': is_multi_table,
                    '_is_cross_db': is_cross_db
                }
                data.append(record)
                
                # 进度显示
                if (idx + 1) % 50 == 0:
                    logger.info(f"  进度: {idx + 1}/{total}")
                
                # 请求间隔
                time.sleep(CONFIG["request_delay"])
            
            logger.info(f"  处理完成: {len(data)} 条QA对")
            return data
            
        except Exception as e:
            logger.error(f"处理文件失败: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def create_excel(self, data: List[Dict], output_path: Path):
        """创建带合并单元格的Excel文件"""
        
        # 排序
        def sort_key(x):
            if x['_is_cross_db']:
                query_type = 2
            elif x['_is_multi_table']:
                query_type = 1
            else:
                query_type = 0
            return (x['所属系统'], x['表名'], query_type)
        
        data_sorted = sorted(data, key=sort_key)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "用表统计"
        
        # 设置列标题（17列）
        headers = [
            '所属系统', '所属数据库', '表名', '判断条件', '查询内容',
            '单表查询问题数量', '单表查询入库问题', '单表查询入库sql',
            '跨表查询问题数量', '跨表查询关联表格', '跨表查询入库问题', '跨表查询入库sql',
            '跨库查询数量', '跨库查询关联数据库', '跨库查询关联表格', '跨库查询入库问题', '跨库查询入库sql'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 设置列宽
        column_widths = [25, 15, 35, 30, 40, 12, 50, 80, 12, 35, 50, 80, 12, 20, 35, 50, 80]
        for i, width in enumerate(column_widths, 1):
            if i <= 26:
                col_letter = chr(64 + i)
            else:
                col_letter = chr(64 + (i-1)//26) + chr(65 + (i-1)%26)
            ws.column_dimensions[col_letter].width = width
        
        current_row = 2
        merge_ranges = []
        
        for (system, table), group in groupby(data_sorted, key=lambda x: (x['所属系统'], x['表名'])):
            group_list = list(group)
            
            single_queries = [r for r in group_list if not r['_is_multi_table'] and not r['_is_cross_db']]
            multi_table_queries = [r for r in group_list if r['_is_multi_table'] and not r['_is_cross_db']]
            cross_db_queries = [r for r in group_list if r['_is_cross_db']]
            
            table_start_row = current_row
            db_name = self.get_database_name(table)
            
            # 写入单表查询
            single_start = current_row
            for i, record in enumerate(single_queries):
                ws.cell(row=current_row, column=1, value=system if current_row == table_start_row else '')
                ws.cell(row=current_row, column=2, value=db_name if current_row == table_start_row else '')
                ws.cell(row=current_row, column=3, value=table if current_row == table_start_row else '')
                ws.cell(row=current_row, column=4, value=record['判断条件'])
                ws.cell(row=current_row, column=5, value=record['查询内容'])
                ws.cell(row=current_row, column=6, value=len(single_queries) if i == 0 else '')
                ws.cell(row=current_row, column=7, value=record['单表查询入库问题'])
                ws.cell(row=current_row, column=8, value=record['单表查询入库sql'])
                current_row += 1
            single_end = current_row - 1 if single_queries else None
            
            # 写入跨表查询
            multi_start = current_row
            for i, record in enumerate(multi_table_queries):
                ws.cell(row=current_row, column=1, value=system if current_row == table_start_row else '')
                ws.cell(row=current_row, column=2, value=db_name if current_row == table_start_row else '')
                ws.cell(row=current_row, column=3, value=table if current_row == table_start_row else '')
                ws.cell(row=current_row, column=4, value=record['判断条件'])
                ws.cell(row=current_row, column=5, value=record['查询内容'])
                ws.cell(row=current_row, column=9, value=len(multi_table_queries) if i == 0 else '')
                ws.cell(row=current_row, column=10, value=record['跨表查询关联表格'])
                ws.cell(row=current_row, column=11, value=record['跨表查询入库问题'])
                ws.cell(row=current_row, column=12, value=record['跨表查询入库sql'])
                current_row += 1
            multi_end = current_row - 1 if multi_table_queries else None
            
            # 写入跨库查询
            cross_start = current_row
            for i, record in enumerate(cross_db_queries):
                ws.cell(row=current_row, column=1, value=system if current_row == table_start_row else '')
                ws.cell(row=current_row, column=2, value=db_name if current_row == table_start_row else '')
                ws.cell(row=current_row, column=3, value=table if current_row == table_start_row else '')
                ws.cell(row=current_row, column=4, value=record['判断条件'])
                ws.cell(row=current_row, column=5, value=record['查询内容'])
                ws.cell(row=current_row, column=13, value=len(cross_db_queries) if i == 0 else '')
                ws.cell(row=current_row, column=14, value=record['跨库查询关联数据库'])
                ws.cell(row=current_row, column=15, value=record['跨库查询关联表格'])
                ws.cell(row=current_row, column=16, value=record['跨库查询入库问题'])
                ws.cell(row=current_row, column=17, value=record['跨库查询入库sql'])
                current_row += 1
            cross_end = current_row - 1 if cross_db_queries else None
            
            table_end_row = current_row - 1
            
            # 合并单元格
            if table_end_row > table_start_row:
                merge_ranges.append((table_start_row, table_end_row, 1, 1))
                merge_ranges.append((table_start_row, table_end_row, 2, 2))
                merge_ranges.append((table_start_row, table_end_row, 3, 3))
            
            if single_queries and single_end and single_end > single_start:
                merge_ranges.append((single_start, single_end, 6, 6))
            if multi_table_queries and multi_end and multi_end > multi_start:
                merge_ranges.append((multi_start, multi_end, 9, 9))
            if cross_db_queries and cross_end and cross_end > cross_start:
                merge_ranges.append((cross_start, cross_end, 13, 13))
        
        # 执行合并
        for start_row, end_row, start_col, end_col in merge_ranges:
            ws.merge_cells(start_row=start_row, start_column=start_col,
                           end_row=end_row, end_column=end_col)
        
        # 设置边框和对齐
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=17):
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border
        
        wb.save(output_path)
        logger.info(f"\n输出文件: {output_path}")
        logger.info(f"总行数: {current_row - 2}")


def main():
    """主函数"""
    print("=" * 70)
    print("QA用表统计生成脚本")
    print("=" * 70)
    
    # 检查Agent ID
    if CONFIG["agent_id"] == "YOUR_AGENT_ID_HERE":
        print("\n错误：请先配置Agent ID！")
        print("步骤：")
        print("1. 在RAGFlow中创建工作流（参见workflow_config.md）")
        print("2. 复制工作流的Agent ID")
        print("3. 将Agent ID填入脚本CONFIG配置中的agent_id字段")
        print("4. 重新运行脚本")
        return
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    # 初始化提取器
    extractor = QAClassificationExtractor(
        base_url=CONFIG["base_url"],
        api_key=CONFIG["api_key"],
        agent_id=CONFIG["agent_id"]
    )
    
    # 初始化生成器
    generator = TableStatisticsGenerator(extractor)
    
    # 处理所有文件
    all_data = []
    
    for file_path, system_key in INPUT_FILES:
        if not Path(file_path).exists():
            logger.warning(f"文件不存在，跳过: {file_path}")
            continue
        
        data = generator.process_qa_file(file_path, system_key)
        all_data.extend(data)
    
    if not all_data:
        logger.error("没有处理到任何数据！")
        return
    
    # 生成Excel
    generator.create_excel(all_data, OUTPUT_FILE)
    
    # 统计信息
    single_count = sum(1 for d in all_data if not d['_is_multi_table'] and not d['_is_cross_db'])
    multi_count = sum(1 for d in all_data if d['_is_multi_table'] and not d['_is_cross_db'])
    cross_count = sum(1 for d in all_data if d['_is_cross_db'])
    
    print("\n" + "=" * 70)
    print("统计汇总")
    print("=" * 70)
    print(f"总QA对数: {len(all_data)}")
    print(f"单表查询: {single_count}")
    print(f"跨表查询: {multi_count}")
    print(f"跨库查询: {cross_count}")
    print("\n处理完成！")


if __name__ == '__main__':
    main()
